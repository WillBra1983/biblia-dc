#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importa traduções PT dos Excel partidos para stepbible_lexicon.sqlite.

Dois modos:

1) Com cabeçalho nas folhas PT (colunas id OU strongs_unified+source, definition_pt, ...)

2) --pair-en-dir  (o teu caso): ficheiros PT só com texto (ex.: coluna A, sem linha de título).
   Junta com os Excel em inglês na mesma pasta de trabalho, mesma numeração parte01…parte10
   e mesma ordem de linhas. As chaves vêm do ficheiro EN (id ou strongs_unified + source).

Dependência:
  pip install openpyxl

Exemplo (PT só texto + EN na área de trabalho):
  python scripts/import_stepbible_pt_from_excel_parts.py ^
    --dir "C:\\Users\\...\\Downloads" --pattern "*_PT.xlsx" ^
    --pair-en-dir "C:\\Users\\...\\Desktop\\StepbibleLexicon_parts" ^
    --dry-run
"""

from __future__ import annotations

import argparse
import re
import sqlite3
from pathlib import Path


def normalizar_unified(s: str) -> str:
    s = str(s or "").strip().upper()
    m = re.match(r"^([HG])0*(\d+)$", s)
    if m:
        return f"{m.group(1)}{int(m.group(2))}"
    return s


def load_workbook_rows(path: Path) -> tuple[list[str], list[list]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    rows = [list(r) for r in rows_iter]
    wb.close()
    return header, rows


def load_pt_single_column(path: Path, col_idx: int, skip_first_rows: int) -> list[str]:
    """Todas as linhas = texto PT (sem cabeçalho), coluna col_idx (0=A)."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    out: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_first_rows:
            continue
        if col_idx >= len(row):
            continue
        v = row[col_idx]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.append(s)
    wb.close()
    return out


def aplicar_map(header: list[str], mapping: dict[str, str]) -> list[str]:
    out = []
    for h in header:
        key = h.strip()
        out.append(mapping.get(key, mapping.get(h, key)))
    return out


def indice_col(header: list[str], nome: str) -> int | None:
    nome_l = nome.strip().lower()
    for i, h in enumerate(header):
        if str(h).strip().lower() == nome_l:
            return i
    return None


def extrair_parte(path: Path) -> tuple[int, int] | None:
    m = re.search(r"parte(\d+)_de_(\d+)", path.name, re.I)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def encontrar_par_ingles(pt_file: Path, en_dir: Path, en_glob: str) -> Path | None:
    """Encontra StepbibleLexicon_parteNN_de_MM.xlsx sem _PT."""
    info = extrair_parte(pt_file)
    if not info:
        return None
    num, total = info
    pad_n = f"{num:02d}"
    pad_t = f"{total:02d}"
    # Aceita qualquer prefixo desde que contenha parteNN_de_MM
    for p in sorted(en_dir.glob(en_glob)):
        if "_PT" in p.name:
            continue
        if re.search(rf"parte{pad_n}_de_{pad_t}", p.name, re.I):
            return p
    return None


def run_pair_mode(
    *,
    db_path: Path,
    pt_dir: Path,
    pt_pattern: str,
    en_dir: Path,
    en_glob: str,
    pt_col: int,
    pt_skip_rows: int,
    dry_run: bool,
) -> int:
    pt_files = sorted(pt_dir.glob(pt_pattern))
    if not pt_pattern.startswith("*"):
        pt_files = sorted(pt_dir.glob(pt_pattern))
    pt_files = [p for p in pt_files if p.is_file()]
    if not pt_files:
        print(f"Nenhum ficheiro em {pt_dir} com padrão {pt_pattern!r}")
        return 1

    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()

    total_ok = 0
    total_miss = 0
    total_skip = 0

    for pt_file in pt_files:
        en_file = encontrar_par_ingles(pt_file, en_dir, en_glob)
        if not en_file:
            print(f"AVISO: sem par em inglês para {pt_file.name} (procurei em {en_dir})")
            continue

        header, en_data = load_workbook_rows(en_file)
        ix_id = indice_col(header, "id")
        ix_u = indice_col(header, "strongs_unified")
        ix_src = indice_col(header, "source")

        if ix_id is None and (ix_u is None or ix_src is None):
            print(f"ERRO: {en_file.name} precisa colunas id OU strongs_unified + source.")
            print(f"  Cabeçalho: {header[:12]}...")
            conn.close()
            return 1

        pt_texts = load_pt_single_column(pt_file, pt_col, pt_skip_rows)
        n = min(len(pt_texts), len(en_data))
        if len(pt_texts) != len(en_data):
            print(
                f"AVISO: {pt_file.name}: {len(pt_texts)} linhas PT vs {len(en_data)} linhas EN — uso {n}."
            )

        for i in range(n):
            row = en_data[i]
            def_pt = pt_texts[i]

            def cell(idx: int | None):
                if idx is None or idx >= len(row):
                    return None
                v = row[idx]
                return None if v is None else str(v).strip()

            if not def_pt.strip():
                total_skip += 1
                continue

            if ix_id is not None:
                try:
                    rid_i = int(row[ix_id])
                except (TypeError, ValueError, IndexError):
                    total_skip += 1
                    continue
                sql = """
                  UPDATE stepbible_lexicon SET
                    definition_pt = ?,
                    definition_clean_pt = COALESCE(definition_clean_pt, ?)
                  WHERE id = ?
                """
                params = (def_pt, def_pt, rid_i)
                check_sql = "SELECT 1 FROM stepbible_lexicon WHERE id = ?"
                check_par = (rid_i,)
            else:
                u = normalizar_unified(cell(ix_u) or "")
                src = cell(ix_src) or ""
                if not u or not src:
                    total_skip += 1
                    continue
                sql = """
                  UPDATE stepbible_lexicon SET
                    definition_pt = ?,
                    definition_clean_pt = COALESCE(definition_clean_pt, ?)
                  WHERE strongs_unified = ? AND source = ?
                """
                params = (def_pt, def_pt, u, src)
                check_sql = "SELECT 1 FROM stepbible_lexicon WHERE strongs_unified = ? AND source = ?"
                check_par = (u, src)

            if dry_run:
                cur.execute(check_sql, check_par)
                if cur.fetchone():
                    total_ok += 1
                else:
                    total_miss += 1
            else:
                cur.execute(sql, params)
                if cur.rowcount == 0:
                    total_miss += 1
                else:
                    total_ok += 1

        print(f"  {pt_file.name} + {en_file.name}: {n} linhas alinhadas")

    if not dry_run:
        conn.commit()
    conn.close()

    print(f"\nAtualizações {'simuladas ' if dry_run else ''}ok: {total_ok}")
    print(f"Sem correspondência no SQLite: {total_miss}")
    print(f"Ignoradas (vazias / sem chave): {total_skip}")
    if dry_run:
        print("(dry-run: base não alterada)")
    else:
        print(f"OK: {db_path}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Importa PT STEPBible de vários Excel para SQLite.")
    ap.add_argument(
        "--db",
        type=Path,
        default=None,
        help="Caminho stepbible_lexicon.sqlite (padrão: public/ no projeto)",
    )
    ap.add_argument("--dir", type=Path, required=True, help="Pasta com os Excel PT")
    ap.add_argument(
        "--pattern",
        type=str,
        default="*_PT.xlsx",
        help="Glob dos ficheiros PT (padrão: *_PT.xlsx)",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Só valida; não grava no SQLite",
    )
    ap.add_argument(
        "--map",
        type=str,
        default="",
        help="Renomes: ColExcel:nome_padrao,... (só modo com cabeçalho no PT)",
    )
    ap.add_argument(
        "--pair-en-dir",
        type=Path,
        default=None,
        help="Pasta com Excel em inglês (mesma parteNN_de_MM, sem _PT). "
        "Modo para PT só com texto na coluna, sem cabeçalho.",
    )
    ap.add_argument(
        "--en-glob",
        type=str,
        default="*.xlsx",
        help="Glob dentro de --pair-en-dir para ficheiros EN (padrão: *.xlsx)",
    )
    ap.add_argument(
        "--pt-text-column",
        type=int,
        default=0,
        help="Índice da coluna com o texto PT no modo pair (0=A). Padrão: 0",
    )
    ap.add_argument(
        "--pt-skip-first-rows",
        type=int,
        default=0,
        help="Saltar N linhas no início do ficheiro PT (padrão: 0)",
    )
    args = ap.parse_args()

    root = Path(__file__).resolve().parents[1]
    db_path = args.db or (root / "public" / "stepbible_lexicon.sqlite")
    if not db_path.exists():
        print(f"SQLite não encontrado: {db_path}")
        return 1

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("Instale: pip install openpyxl")
        return 1

    if args.pair_en_dir is not None:
        en_dir = args.pair_en_dir.resolve()
        if not en_dir.is_dir():
            print(f"Pasta EN não existe: {en_dir}")
            return 1
        return run_pair_mode(
            db_path=db_path,
            pt_dir=args.dir.resolve(),
            pt_pattern=args.pattern,
            en_dir=en_dir,
            en_glob=args.en_glob,
            pt_col=args.pt_text_column,
            pt_skip_rows=args.pt_skip_first_rows,
            dry_run=args.dry_run,
        )

    mapping: dict[str, str] = {}
    if args.map.strip():
        for part in args.map.split(","):
            part = part.strip()
            if ":" in part:
                k, v = part.split(":", 1)
                mapping[k.strip()] = v.strip()

    files = sorted(args.dir.glob(args.pattern))
    if not files:
        print(f"Nenhum ficheiro com padrão {args.pattern!r} em {args.dir}")
        print(
            "Dica: se o PT não tem linha de cabeçalho, use:\n"
            '  --pair-en-dir "C:\\...\\StepbibleLexicon_parts"\n'
            "(pasta dos Excel em inglês com id ou strongs_unified + source)."
        )
        return 1

    todas_linhas: list[tuple] = []
    for fp in files:
        header, rows = load_workbook_rows(fp)
        header = aplicar_map(header, mapping)
        todas_linhas.extend([(fp.name, header, r) for r in rows if any(x is not None and str(x).strip() for x in r)])

    if not todas_linhas:
        print("Nenhuma linha de dados.")
        return 1

    header0 = todas_linhas[0][1]
    ix_u = indice_col(header0, "strongs_unified")
    ix_src = indice_col(header0, "source")
    ix_id = indice_col(header0, "id")
    ix_dpt = indice_col(header0, "definition_pt")
    ix_dcpt = indice_col(header0, "definition_clean_pt")
    ix_gpt = indice_col(header0, "gloss_pt")

    if ix_id is None and (ix_u is None or ix_src is None):
        print(
            "Cabeçalho precisa de 'id' OU ('strongs_unified' e 'source').\n"
            f"Cabeçalho encontrado: {header0[:3]}...\n\n"
            "Se os ficheiros PT são só texto (sem títulos de coluna), use:\n"
            '  --pair-en-dir "C:\\Users\\...\\Desktop\\StepbibleLexicon_parts"'
        )
        return 1
    if ix_dpt is None and ix_dcpt is None:
        print("Precisa da coluna definition_pt ou definition_clean_pt.")
        return 1

    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()

    atualizados = 0
    nao_achados = 0
    vazios = 0

    for fname, header, row in todas_linhas:

        def cell(i: int | None):
            if i is None or i >= len(row):
                return None
            v = row[i]
            return None if v is None else str(v).strip()

        def_pt = cell(ix_dpt) if ix_dpt is not None else ""
        dc_pt = cell(ix_dcpt) if ix_dcpt is not None else ""
        g_pt = cell(ix_gpt) if ix_gpt is not None else ""

        if not def_pt and not dc_pt:
            vazios += 1
            continue

        if ix_id is not None:
            rid = row[ix_id]
            try:
                rid_i = int(rid)
            except (TypeError, ValueError):
                vazios += 1
                continue
            sql = """
              UPDATE stepbible_lexicon SET
                definition_pt = COALESCE(?, definition_pt),
                definition_clean_pt = COALESCE(?, definition_clean_pt),
                gloss_pt = COALESCE(?, gloss_pt)
              WHERE id = ?
            """
            params = (
                def_pt or None,
                dc_pt or None,
                g_pt or None,
                rid_i,
            )
            rid_i_use = rid_i
            u_use = None
            src_use = None
        else:
            u = normalizar_unified(cell(ix_u) or "")
            src = cell(ix_src) or ""
            if not u or not src:
                vazios += 1
                continue
            sql = """
              UPDATE stepbible_lexicon SET
                definition_pt = COALESCE(?, definition_pt),
                definition_clean_pt = COALESCE(?, definition_clean_pt),
                gloss_pt = COALESCE(?, gloss_pt)
              WHERE strongs_unified = ? AND source = ?
            """
            params = (def_pt or None, dc_pt or None, g_pt or None, u, src)
            rid_i_use = None
            u_use = u
            src_use = src

        if args.dry_run:
            if rid_i_use is not None:
                cur.execute("SELECT 1 FROM stepbible_lexicon WHERE id = ?", (rid_i_use,))
            else:
                cur.execute(
                    "SELECT 1 FROM stepbible_lexicon WHERE strongs_unified = ? AND source = ?",
                    (u_use, src_use),
                )
            if cur.fetchone():
                atualizados += 1
            else:
                nao_achados += 1
            continue

        cur.execute(sql, params)
        if cur.rowcount == 0:
            nao_achados += 1
        else:
            atualizados += 1

    if not args.dry_run:
        conn.commit()
    conn.close()

    print(f"Ficheiros: {len(files)}")
    print(f"Linhas processadas (com texto PT): {atualizados}")
    print(f"Sem correspondência no SQLite: {nao_achados}")
    print(f"Ignoradas (vazias / sem chave): {vazios}")
    if args.dry_run:
        print("(dry-run: base não foi alterada)")
    else:
        print(f"OK: {db_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
