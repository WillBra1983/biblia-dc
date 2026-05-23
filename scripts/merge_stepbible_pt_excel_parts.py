#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Junta vários Excel PT (só texto numa coluna, sem linha de cabeçalho) por ordem parte01…parte10.

Saída:
  - merged_stepbible_pt.xlsx  (uma coluna A, todas as linhas na ordem global)
  - merged_stepbible_pt.txt   (opcional: --txt, separador ---STEPBIBLE-ENTRY---)

Dependência: pip install openpyxl

Uso:
  python scripts/merge_stepbible_pt_excel_parts.py --dir "C:\\...\\Downloads" --pattern "*_PT.xlsx"
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path


def extrair_ordem_parte(path: Path) -> tuple[int, int, str]:
    """Ordena por (total_de, numero_parte, nome) para ficar 01,02,...,10 dentro do mesmo conjunto."""
    m = re.search(r"parte(\d+)_de_(\d+)", path.name, re.I)
    if m:
        num = int(m.group(1))
        total = int(m.group(2))
        return (total, num, path.name.lower())
    return (999, 999, path.name.lower())


def ler_coluna_sem_cabecalho(path: Path, col_idx: int) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    out: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if col_idx >= len(row):
            continue
        v = row[col_idx]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.append(s)
    wb.close()
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Junta partes Excel PT num só ficheiro.")
    ap.add_argument("--dir", type=Path, required=True, help="Pasta com os *_PT.xlsx")
    ap.add_argument("--pattern", type=str, default="*_PT.xlsx", help="Glob (padrão: *_PT.xlsx)")
    ap.add_argument("--column", type=int, default=0, help="Índice da coluna (0=A)")
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Ficheiro Excel de saída (padrão: merged_stepbible_pt.xlsx na pasta --dir)",
    )
    ap.add_argument(
        "--txt",
        action="store_true",
        help="Também grava .txt com separador ---STEPBIBLE-ENTRY--- entre entradas",
    )
    args = ap.parse_args()

    d = args.dir.resolve()
    if not d.is_dir():
        print(f"Pasta não existe: {d}")
        return 1

    try:
        from openpyxl import Workbook
    except ImportError:
        print("Instale: pip install openpyxl")
        return 1

    files = sorted(d.glob(args.pattern), key=extrair_ordem_parte)
    files = [p for p in files if p.is_file()]
    if not files:
        print(f"Nenhum ficheiro com {args.pattern!r} em {d}")
        return 1

    todas_linhas: list[str] = []
    for fp in files:
        chunk = ler_coluna_sem_cabecalho(fp, args.column)
        todas_linhas.extend(chunk)
        print(f"  {fp.name}: {len(chunk)} linha(s)")

    out_xlsx = args.out
    if out_xlsx is None:
        out_xlsx = d / "merged_stepbible_pt.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "PT"
    for i, texto in enumerate(todas_linhas, start=1):
        ws.cell(row=i, column=1, value=texto)
    wb.save(str(out_xlsx))
    print(f"\nExcel: {out_xlsx} ({len(todas_linhas)} linhas)")

    if args.txt:
        out_txt = out_xlsx.with_suffix(".txt")
        sep = "\n---STEPBIBLE-ENTRY---\n"
        out_txt.write_text(sep.join(todas_linhas), encoding="utf-8")
        print(f"TXT:   {out_txt} (separador entre entradas: ---STEPBIBLE-ENTRY---)")

    print(
        "\nPara colar no SQLite: importe esta coluna alinhada por ordem com os ids na base,"
        "\nou use o script import_stepbible_pt_from_excel_parts.py com --pair-en-dir dos Excel EN."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
