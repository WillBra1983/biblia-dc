#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Divide uma folha de um ficheiro Excel (.xlsx) em N partes (por padrão 10),
repetindo a primeira linha (cabeçalho) em cada parte.

Dependência:
  pip install openpyxl

Uso:
  python scripts/split_excel_into_parts.py planilha.xlsx
  python scripts/split_excel_into_parts.py planilha.xlsx --parts 10 --sheet 0
"""

from __future__ import annotations

import argparse
from pathlib import Path


def main() -> int:
    ap = argparse.ArgumentParser(description="Divide um Excel em várias partes (por linhas).")
    ap.add_argument("xlsx", type=Path, help="Caminho do ficheiro .xlsx")
    ap.add_argument(
        "--parts",
        type=int,
        default=10,
        metavar="N",
        help="Número de partes (padrão: 10)",
    )
    ap.add_argument(
        "--sheet",
        type=str,
        default="0",
        help="Nome da folha ou índice (0 = primeira). Padrão: 0",
    )
    ap.add_argument(
        "--out-dir",
        type=Path,
        default=None,
        help="Pasta de saída (padrão: subpasta nome_arquivo_parts)",
    )
    ap.add_argument(
        "--prefix",
        type=str,
        default=None,
        help="Prefixo dos ficheiros (padrão: nome base do xlsx)",
    )
    ap.add_argument(
        "--no-header",
        action="store_true",
        help="Não repetir a 1.ª linha como cabeçalho em cada parte",
    )
    args = ap.parse_args()

    src = args.xlsx.resolve()
    if not src.exists():
        print(f"Ficheiro não encontrado: {src}")
        return 1
    if src.suffix.lower() not in (".xlsx",):
        print("Use ficheiro .xlsx (formato Office Open XML).")
        return 1

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Instale: pip install openpyxl")
        return 1

    parts = max(1, int(args.parts))
    wb = load_workbook(str(src), read_only=True, data_only=True)
    try:
        if args.sheet.isdigit():
            idx = int(args.sheet)
            if idx < 0 or idx >= len(wb.sheetnames):
                print(f"Índice de folha inválido. Folhas: {wb.sheetnames}")
                return 1
            sheet_name = wb.sheetnames[idx]
        else:
            if args.sheet not in wb.sheetnames:
                print(f"Folha '{args.sheet}' não existe. Disponíveis: {wb.sheetnames}")
                return 1
            sheet_name = args.sheet
    finally:
        wb.close()

    # Reabre em modo normal para copiar células (read_only não permite write)
    wb = load_workbook(str(src), data_only=True)
    ws = wb[sheet_name]
    all_rows: list[tuple] = []
    for row in ws.iter_rows():
        all_rows.append(tuple(cell.value for cell in row))

    if not all_rows:
        print("Folha vazia.")
        wb.close()
        return 1

    header = all_rows[0] if not args.no_header else None
    data_rows = all_rows[1:] if not args.no_header else all_rows
    n = len(data_rows)
    if n == 0:
        print("Sem linhas de dados (só cabeçalho ou vazio).")
        wb.close()
        return 1

    out_dir = args.out_dir
    if out_dir is None:
        out_dir = src.parent / f"{src.stem}_parts"
    out_dir = Path(out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    prefix = args.prefix or src.stem

    base = n // parts
    extra = n % parts
    start = 0
    part_num = 0
    for i in range(parts):
        size = base + (1 if i < extra else 0)
        if size <= 0:
            continue
        chunk = data_rows[start : start + size]
        start += size
        part_num += 1

        from openpyxl import Workbook

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = ws.title[:31]  # limite Excel

        row_out = 1
        if header is not None:
            for col, val in enumerate(header, start=1):
                out_ws.cell(row=row_out, column=col, value=val)
            row_out += 1
        for data_row in chunk:
            for col, val in enumerate(data_row, start=1):
                out_ws.cell(row=row_out, column=col, value=val)
            row_out += 1

        out_path = out_dir / f"{prefix}_parte{part_num:02d}_de_{parts:02d}.xlsx"
        out_wb.save(str(out_path))
        print(f"{out_path.name}: {len(chunk)} linha(s) de dados")

    wb.close()
    print(f"\nTotal: {n} linha(s) de dados em {part_num} ficheiro(s) -> {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
